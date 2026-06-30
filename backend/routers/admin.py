import random
import io
from datetime import datetime, timedelta, timezone
from fastapi import APIRouter, Depends, HTTPException, Request as FastAPIRequest
from fastapi.responses import Response
from slowapi import Limiter
from slowapi.util import get_remote_address
from sqlalchemy.orm import Session

import models
import schemas
import auth
from database import get_db
from utils import send_gmail, _admin_codes

router = APIRouter(prefix="/api/admin", tags=["admin"])
limiter = Limiter(key_func=get_remote_address)


@router.post("/verify-2fa")
@limiter.limit("5/minute")
def verify_admin_2fa(request: FastAPIRequest, data: schemas.AdminVerifyCode, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == data.email).first()
    if not user or not user.is_admin:
        raise HTTPException(status_code=401, detail="Доступ запрещён")

    stored_code = getattr(user, 'admin_code', None)
    expire_time = getattr(user, 'admin_code_expires', None)
    code_source = "db"

    if stored_code is None and data.email in _admin_codes:
        fb = _admin_codes[data.email]
        stored_code = fb["code"]
        expire_time = fb["expires"]
        code_source = "memory"

    if not stored_code or stored_code != data.code:
        raise HTTPException(status_code=401, detail="Неверный код подтверждения")

    if expire_time:
        if expire_time.tzinfo is None:
            expire_time = expire_time.replace(tzinfo=timezone.utc)
        if expire_time < datetime.now(timezone.utc):
            raise HTTPException(status_code=401, detail="Код просрочен. Запросите новый.")

    if code_source == "db":
        try:
            user.admin_code = None
            user.admin_code_expires = None
            db.commit()
        except Exception:
            db.rollback()
    else:
        _admin_codes.pop(data.email, None)

    token = auth.create_access_token({"sub": str(user.id)})
    return {
        "access_token": token,
        "token_type": "bearer",
        "user_name": user.name,
        "is_admin": True
    }


@router.post("/send-2fa-code")
def resend_admin_2fa(data: schemas.AdminVerifyCode, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == data.email).first()
    if not user or not user.is_admin:
        raise HTTPException(status_code=401, detail="Доступ запрещён")

    admin_code = str(random.randint(1000, 9999))
    code_saved = False
    try:
        user.admin_code = admin_code
        user.admin_code_expires = datetime.now(timezone.utc) + timedelta(minutes=10)
        db.commit()
        code_saved = True
    except Exception:
        db.rollback()

    if not code_saved:
        _admin_codes[user.email] = {
            "code": admin_code,
            "expires": datetime.now(timezone.utc) + timedelta(minutes=10)
        }

    send_gmail(
        to=user.email,
        subject="Новый код подтверждения — Minsk Gastro Guide",
        body=(
            f"Здравствуйте, {user.name}!\n\n"
            f"Ваш новый код для входа: {admin_code}\n\n"
            f"Код действителен 10 минут."
        )
    )

    return {"message": "Код отправлен на почту"}


@router.get("/bookings")
def get_all_bookings(
    page: int = 1,
    per_page: int = 50,
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(auth.get_current_admin)
):
    query = db.query(models.Booking)
    total = query.count()
    bookings = query.order_by(models.Booking.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()
    return {
        "items": [
            {
                "id": b.id,
                "venue_id": b.venue_id,
                "venue_name": b.venue_name,
                "name": b.name,
                "date": str(b.date),
                "guests": b.guests,
                "phone": b.phone,
                "message": b.message,
                "cancel_reason": b.cancel_reason,
                "status": b.status,
                "created_at": str(b.created_at)
            }
            for b in bookings
        ],
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page
    }


@router.patch("/bookings/{booking_id}/status")
def update_booking_status(
    booking_id: int,
    data: schemas.AdminBookingStatus,
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(auth.get_current_admin)
):
    booking = db.query(models.Booking).filter(models.Booking.id == booking_id).first()
    if not booking:
        raise HTTPException(status_code=404, detail="Бронирование не найдено")

    booking.status = data.status
    if data.reason:
        booking.cancel_reason = data.reason
    db.commit()
    return {"message": "Статус обновлён", "id": booking.id}


@router.get("/messages")
def get_all_messages(
    page: int = 1,
    per_page: int = 50,
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(auth.get_current_admin)
):
    query = db.query(models.Message)
    total = query.count()
    messages = query.order_by(models.Message.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()
    return {
        "items": [
            {
                "id": m.id,
                "name": m.name,
                "email": m.email,
                "subject": m.subject,
                "message": m.message,
                "source": m.source,
                "category": m.category,
                "created_at": str(m.created_at)
            }
            for m in messages
        ],
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page
    }


@router.get("/export/bookings")
def export_bookings_csv(
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(auth.get_current_admin)
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    bookings = db.query(models.Booking).order_by(models.Booking.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Бронирования"

    headers = ["ID", "Имя", "Телефон", "Заведение", "Филиал (ID)", "Дата", "Гости", "Пожелания", "Статус", "Причина отмены", "Дата создания"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    active_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    cancelled_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    status_font_active = Font(color="2E7D32", bold=True)
    status_font_cancelled = Font(color="C62828", bold=True)

    for row_idx, b in enumerate(bookings, 2):
        status_display = "Активно" if b.status == "active" else "Отменено"
        created_display = b.created_at.strftime("%d.%m.%Y %H:%M") if b.created_at else ""

        values = [b.id, b.name, b.phone, b.venue_name, b.venue_id, str(b.date), b.guests, b.message or "", status_display, b.cancel_reason or "", created_display]

        row_fill = active_fill if b.status == "active" else cancelled_fill

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col in [8, 10]))

            if col == 9:
                cell.fill = row_fill
                cell.font = status_font_active if b.status == "active" else status_font_cancelled

    col_widths = [6, 20, 18, 25, 15, 14, 8, 30, 12, 25, 18]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bookings_report.xlsx"}
    )


@router.get("/export/messages")
def export_messages_csv(
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(auth.get_current_admin)
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    messages = db.query(models.Message).order_by(models.Message.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Обращения"

    headers = ["ID", "Имя", "Email", "Тема", "Сообщение", "Источник", "Категория", "Дата"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    contact_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    support_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    source_font_contact = Font(color="1565C0", bold=True)
    source_font_support = Font(color="E65100", bold=True)

    for row_idx, m in enumerate(messages, 2):
        source_display = "Контакты" if m.source == "contact" else "Поддержка"
        created_display = m.created_at.strftime("%d.%m.%Y %H:%M") if m.created_at else ""

        values = [m.id, m.name, m.email, m.subject, m.message, source_display, m.category or "", created_display]

        row_fill = contact_fill if m.source == "contact" else support_fill

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col in [4, 5]))

            if col == 6:
                cell.fill = row_fill
                cell.font = source_font_contact if m.source == "contact" else source_font_support

    col_widths = [6, 20, 25, 30, 40, 14, 18, 18]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=messages_report.xlsx"}
    )
