import os
import random
import json
from datetime import datetime, timedelta, timezone
from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware # <--- ИМПОРТИРУЕМ CORS
from sqlalchemy.orm import Session
from database import engine, Base, get_db
import models
import schemas
import auth
import resend

# Создаем таблицы
Base.metadata.create_all(bind=engine)

from sqlalchemy import text
try:
    with engine.connect() as conn:
        conn.execute(text("ALTER TABLE reviews ADD COLUMN IF NOT EXISTS photos TEXT"))
        conn.execute(text("ALTER TABLE reviews ADD COLUMN IF NOT EXISTS branch_id VARCHAR(50)"))
        try:
            conn.execute(text("ALTER TABLE reviews DROP COLUMN IF EXISTS branch_address"))
        except Exception:
            pass
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS is_admin BOOLEAN DEFAULT FALSE"))
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS admin_code VARCHAR(10)"))
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS admin_code_expires TIMESTAMP WITH TIME ZONE"))
        conn.commit()
except Exception as e:
    print(f"Migration warning: {e}")

# In-memory fallback для 2FA кодов (если колонки admin_code не существуют в БД)
_admin_codes = {}  # email -> {"code": "1234", "expires": datetime}


app = FastAPI(
    title="Minsk Gastro Guide API",
    description="Backend для путеводителя по заведениям Минска",
    version="1.0.0"
)

# ==========================================
# НАСТРОЙКА CORS (ОТКРЫВАЕМ ДВЕРИ ДЛЯ ФРОНТЕНДА)
# ==========================================
app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://praktika2-eta.vercel.app"],
    allow_credentials=True,
    allow_methods=["*"],  # Разрешаем GET, POST, OPTIONS и т.д.
    allow_headers=["*"],  # Разрешаем любые заголовки (включая токены)
)

@app.get("/")
def read_root():
    return {"status": "success", "message": "API работает!"}

@app.post("/api/register", response_model=schemas.UserResponse)
def register_user(user: schemas.UserCreate, db: Session = Depends(get_db)):
    existing_user = db.query(models.User).filter(models.User.email == user.email).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="Пользователь с таким email уже существует")
    
    hashed_pwd = auth.hash_password(user.password)
    new_user = models.User(name=user.name, email=user.email, password_hash=hashed_pwd, phone=user.phone)
    
    db.add(new_user)
    db.commit()
    db.refresh(new_user) 
    
    return new_user 

@app.post("/api/login")
def login_user(user: schemas.UserLogin, db: Session = Depends(get_db)):
    db_user = db.query(models.User).filter(models.User.email == user.email).first()
    if not db_user or not auth.verify_password(user.password, db_user.password_hash):
        raise HTTPException(status_code=401, detail="Неверный email или пароль")
    
    # Если пользователь — администратор, отправляем 2FA код
    if db_user.is_admin:
        admin_code = str(random.randint(1000, 9999))
        code_saved = False
        try:
            db_user.admin_code = admin_code
            db_user.admin_code_expires = datetime.now(timezone.utc) + timedelta(minutes=10)
            db.commit()
            code_saved = True
        except Exception:
            db.rollback()

        if not code_saved:
            _admin_codes[db_user.email] = {
                "code": admin_code,
                "expires": datetime.now(timezone.utc) + timedelta(minutes=10)
            }

        resend_api_key = os.getenv("RESEND_API_KEY")
        if resend_api_key and resend_api_key != "re_СЮДА_ВСТАВЬ_СВОЙ_КЛЮЧ":
            try:
                resend.api_key = resend_api_key
                resend.Emails.send({
                    "from": os.getenv("RESEND_FROM_EMAIL", "Minsk Gastro Guide <onboarding@resend.dev>"),
                    "to": [db_user.email],
                    "subject": "Код подтверждения входа — Minsk Gastro Guide",
                    "text": (
                        f"Здравствуйте, {db_user.name}!\n\n"
                        f"Код для входа в панель администратора: {admin_code}\n\n"
                        f"Код действителен 10 минут. Если вы не запрашивали вход — проигнорируйте это письмо."
                    )
                })
            except Exception as e:
                print(f"Ошибка отправки 2FA кода: {e}")

        return {"requires_2fa": True, "email": db_user.email}
    
    token = auth.create_access_token({"sub": str(db_user.id)})
    return {
        "access_token": token,
        "token_type": "bearer",
        "user_name": db_user.name,
        "is_admin": False
    }
    
    # === НОВЫЙ МАРШРУТ: ПОЛУЧЕНИЕ ПРОФИЛЯ ===
# Обрати внимание на Depends(auth.get_current_user) - сюда нельзя без токена!
@app.get("/api/users/me", response_model=schemas.UserResponse)
def get_user_profile(current_user: models.User = Depends(auth.get_current_user)):
    # Если функция дошла сюда, значит охранник пропустил нас. 
    # В current_user уже лежат данные того, чей токен мы прислали.
    return current_user

@app.post("/api/favorites/toggle", response_model=schemas.FavoriteResponse)
def toggle_favorite(
    favorite: schemas.FavoriteCreate, 
    db: Session = Depends(get_db), 
    current_user: models.User = Depends(auth.get_current_user) # Защита: только для авторизованных
):
    # Ищем, есть ли уже такой лайк от этого юзера на это заведение
    existing_fav = db.query(models.Favorite).filter(
        models.Favorite.user_id == current_user.id,
        models.Favorite.venue_id == favorite.venue_id
    ).first()

    if existing_fav:
        # Если лайк найден — удаляем его (Toggle OFF)
        db.delete(existing_fav)
        db.commit()
        return {"venue_id": favorite.venue_id, "message": "Удалено из избранного"}
    else:
        # Если лайка нет — создаем новый (Toggle ON)
        new_fav = models.Favorite(user_id=current_user.id, venue_id=favorite.venue_id)
        db.add(new_fav)
        db.commit()
        return {"venue_id": favorite.venue_id, "message": "Добавлено в избранное"}


@app.get("/api/favorites")
def get_user_favorites(
    db: Session = Depends(get_db), 
    current_user: models.User = Depends(auth.get_current_user)
):
    # Находим все лайки этого пользователя
    favorites = db.query(models.Favorite).filter(models.Favorite.user_id == current_user.id).all()
    
    # Извлекаем из объектов БД только список строк (ID заведений)
    # Получится массив вроде: ['zerno', 'lidbeer']
    venue_ids = [fav.venue_id for fav in favorites]
    return venue_ids


# ==========================================
# УДАЛЕНИЕ И РЕДАКТИРОВАНИЕ АККАУНТА
# ==========================================

@app.patch("/api/users/me", response_model=schemas.UserResponse)
def update_user_profile(
    user_update: schemas.UserUpdate, 
    db: Session = Depends(get_db), 
    current_user: models.User = Depends(auth.get_current_user)
):
    if user_update.name is not None:
        current_user.name = user_update.name
    if user_update.phone is not None:
        current_user.phone = user_update.phone
        
    db.commit()
    db.refresh(current_user)
    return current_user

@app.delete("/api/users/me")
def delete_user_profile(
    db: Session = Depends(get_db), 
    current_user: models.User = Depends(auth.get_current_user)
):
    # Благодаря ondelete="CASCADE" в моделях, удаление юзера удалит и его лайки/брони
    db.delete(current_user)
    db.commit()
    return {"message": "Аккаунт успешно удален"}

# ==========================================
# АВАТАР ПОЛЬЗОВАТЕЛЯ
# ==========================================

@app.post("/api/users/avatar", response_model=schemas.UserResponse)
async def upload_avatar(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    if file.content_type not in ["image/jpeg", "image/png", "image/webp", "image/gif"]:
        raise HTTPException(status_code=400, detail="Допустимые форматы: JPEG, PNG, WebP, GIF")

    import base64
    contents = await file.read()
    if len(contents) > 2 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Максимальный размер файла — 2 МБ")

    b64 = base64.b64encode(contents).decode("utf-8")
    current_user.avatar = f"data:{file.content_type};base64,{b64}"
    db.commit()
    db.refresh(current_user)
    return current_user

@app.delete("/api/users/avatar", response_model=schemas.UserResponse)
def delete_avatar(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    current_user.avatar = None
    db.commit()
    db.refresh(current_user)
    return current_user

# ==========================================
# СМЕНА ПАРОЛЯ ИЗ ЛИЧНОГО КАБИНЕТА
# ==========================================

@app.post("/api/send-change-code")
def send_change_code(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    reset_code = str(random.randint(1000, 9999))
    current_user.reset_code = reset_code
    current_user.reset_code_expires = datetime.now(timezone.utc) + timedelta(minutes=15)
    db.commit()

    resend_api_key = os.getenv("RESEND_API_KEY")
    if resend_api_key and resend_api_key != "re_СЮДА_ВСТАВЬ_СВОЙ_КЛЮЧ":
        try:
            resend.api_key = resend_api_key
            resend.Emails.send({
                "from": os.getenv("RESEND_FROM_EMAIL", "Minsk Gastro Guide <onboarding@resend.dev>"),
                "to": [current_user.email],
                "subject": "Смена пароля - Minsk Gastro Guide",
                "text": (
                    f"Здравствуйте, {current_user.name}!\n\n"
                    f"Вы запросили смену пароля на сайте Minsk Gastro Guide.\n"
                    f"Ваш код подтверждения: {reset_code}\n\n"
                    f"Код действителен 15 минут. Если вы не запрашивали смену пароля — проигнорируйте это письмо."
                )
            })
        except Exception as e:
            print(f"Ошибка отправки письма: {e}")
            raise HTTPException(status_code=500, detail="Ошибка почтового сервера")

    return {"message": "Код подтверждения отправлен на почту"}

@app.post("/api/change-password")
def change_password(
    data: schemas.ChangePassword,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    if not current_user.reset_code or current_user.reset_code != data.code:
        raise HTTPException(status_code=400, detail="Неверный код подтверждения")

    expire_time = current_user.reset_code_expires
    if expire_time.tzinfo is None:
        expire_time = expire_time.replace(tzinfo=timezone.utc)

    if expire_time < datetime.now(timezone.utc):
        raise HTTPException(status_code=400, detail="Код просрочен")

    current_user.password_hash = auth.hash_password(data.new_password)
    current_user.reset_code = None
    current_user.reset_code_expires = None
    db.commit()

    return {"message": "Пароль успешно изменен"}

# ==========================================
# ВОССТАНОВЛЕНИЕ ПАРОЛЯ С ОТПРАВКОЙ НА GMAIL
# ==========================================

@app.post("/api/forgot-password")
def forgot_password(data: schemas.ForgotPassword, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == data.email).first()
    if not user:
        # Для безопасности мы не говорим, что email не найден, просто возвращаем ОК
        return {"message": "Если email существует, код отправлен"}

    # 1. Генерируем 4-значный код
    reset_code = str(random.randint(1000, 9999))
    user.reset_code = reset_code
    user.reset_code_expires = datetime.now(timezone.utc) + timedelta(minutes=15)
    db.commit()

    # 2. Отправляем письмо через Resend
    resend_api_key = os.getenv("RESEND_API_KEY")
    
    if resend_api_key and resend_api_key != "re_СЮДА_ВСТАВЬ_СВОЙ_КЛЮЧ":
        try:
            resend.api_key = resend_api_key
            resend.Emails.send({
                "from": os.getenv("RESEND_FROM_EMAIL", "Minsk Gastro Guide <onboarding@resend.dev>"),
                "to": [user.email],
                "subject": "Восстановление пароля - Minsk Gastro Guide",
                "text": f"Здравствуйте, {user.name}!\n\nВаш код для восстановления пароля на сайте Minsk Gastro Guide: {reset_code}\n\nКод действителен 15 минут."
            })
        except Exception as e:
            print(f"Ошибка отправки письма: {e}")
            raise HTTPException(status_code=500, detail="Ошибка почтового сервера")

    return {"message": "Код отправлен на почту"}


@app.post("/api/reset-password")
def reset_password(data: schemas.ResetPassword, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == data.email).first()
    
    expire_time = user.reset_code_expires
    if expire_time.tzinfo is None:
        expire_time = expire_time.replace(tzinfo=timezone.utc)

    # Сравниваем два таймзон-зависимых времени
    if expire_time < datetime.now(timezone.utc):
        raise HTTPException(status_code=400, detail="Код просрочен")
    
    # Шифруем новый пароль и стираем код
    user.password_hash = auth.hash_password(data.new_password)
    user.reset_code = None
    user.reset_code_expires = None
    db.commit()

    return {"message": "Пароль успешно изменен"}

# ==========================================
# БРОНИРОВАНИЯ
# ==========================================

@app.post("/api/bookings", response_model=schemas.BookingResponse)
def create_booking(
    booking: schemas.BookingCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    new_booking = models.Booking(
        venue_id=booking.venue_id,
        venue_name=booking.venue_name,
        name=booking.name,
        date=booking.date,
        guests=booking.guests,
        phone=booking.phone,
        message=booking.message,
        user_id=current_user.id,
        status="active"
    )
    db.add(new_booking)
    db.commit()
    db.refresh(new_booking)

    return schemas.BookingResponse(
        id=new_booking.id,
        venue_id=new_booking.venue_id,
        venue_name=new_booking.venue_name,
        name=new_booking.name,
        date=new_booking.date,
        guests=new_booking.guests,
        phone=new_booking.phone,
        message=new_booking.message,
        cancel_reason=new_booking.cancel_reason,
        status=new_booking.status,
        created_at=str(new_booking.created_at)
    )

@app.get("/api/bookings", response_model=list[schemas.BookingResponse])
def get_user_bookings(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    bookings = db.query(models.Booking).filter(
        models.Booking.user_id == current_user.id
    ).order_by(models.Booking.created_at.desc()).all()

    return [
        schemas.BookingResponse(
            id=b.id,
            venue_id=b.venue_id,
            venue_name=b.venue_name,
            name=b.name,
            date=b.date,
            guests=b.guests,
            phone=b.phone,
            message=b.message,
            cancel_reason=b.cancel_reason,
            status=b.status,
            created_at=str(b.created_at)
        )
        for b in bookings
    ]

@app.patch("/api/bookings/{booking_id}/cancel")
def cancel_booking(
    booking_id: int,
    data: schemas.CancelBooking,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    booking = db.query(models.Booking).filter(
        models.Booking.id == booking_id,
        models.Booking.user_id == current_user.id
    ).first()

    if not booking:
        raise HTTPException(status_code=404, detail="Бронирование не найдено")

    if booking.status == "cancelled":
        raise HTTPException(status_code=400, detail="Бронирование уже отменено")

    booking.status = "cancelled"
    booking.cancel_reason = data.reason
    db.commit()
    return {"message": "Бронирование отменено", "id": booking.id}

# ==========================================
# ОТЗЫВЫ
# ==========================================

@app.post("/api/reviews", response_model=schemas.ReviewResponse)
def create_review(
    review: schemas.ReviewCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    existing = db.query(models.Review).filter(
        models.Review.user_id == current_user.id,
        models.Review.venue_id == review.venue_id,
        models.Review.branch_id == review.branch_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Вы уже оставляли отзыв на этот адрес")

    new_review = models.Review(
        rating=review.rating,
        text=review.text,
        venue_id=review.venue_id,
        branch_id=review.branch_id,
        user_id=current_user.id
    )
    db.add(new_review)
    db.commit()
    db.refresh(new_review)

    return schemas.ReviewResponse(
        id=new_review.id,
        rating=new_review.rating,
        text=new_review.text,
        photos=None,
        venue_id=new_review.venue_id,
        branch_id=new_review.branch_id,
        created_at=str(new_review.created_at),
        user_name=current_user.name,
        user_avatar=current_user.avatar,
        user_id=current_user.id
    )


@app.post("/api/reviews/photos")
async def upload_review_photos(
    review_id: int = None,
    files: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    import base64
    if len(files) > 3:
        raise HTTPException(status_code=400, detail="Максимум 3 фотографии")

    photos = []
    for file in files:
        if file.content_type not in ["image/jpeg", "image/png", "image/webp", "image/gif"]:
            raise HTTPException(status_code=400, detail="Допустимые форматы: JPEG, PNG, WebP, GIF")
        contents = await file.read()
        if len(contents) > 5 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="Максимальный размер файла — 5 МБ")
        b64 = base64.b64encode(contents).decode("utf-8")
        photos.append(f"data:{file.content_type};base64,{b64}")

    if review_id is not None:
        review = db.query(models.Review).filter(
            models.Review.id == review_id,
            models.Review.user_id == current_user.id
        ).first()
        if review:
            review.photos = json.dumps(photos)
            db.commit()

    return {"photos": photos}


@app.get("/api/reviews/my", response_model=list[schemas.ReviewResponse])
def get_my_reviews(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    reviews = db.query(models.Review).filter(
        models.Review.user_id == current_user.id
    ).order_by(models.Review.created_at.desc()).all()

    result = []
    for r in reviews:
        user = r.user
        photos = json.loads(r.photos) if r.photos else None
        result.append(schemas.ReviewResponse(
            id=r.id,
            rating=r.rating,
            text=r.text,
            photos=photos,
            venue_id=r.venue_id,
            branch_id=r.branch_id,
            created_at=str(r.created_at),
            user_name=user.name if user else "Удалённый пользователь",
            user_avatar=user.avatar if user else None,
            user_id=r.user_id
        ))
    return result


@app.get("/api/reviews/{venue_id}", response_model=list[schemas.ReviewResponse])
def get_venue_reviews(venue_id: str, db: Session = Depends(get_db)):
    reviews = db.query(models.Review).filter(
        models.Review.venue_id == venue_id
    ).order_by(models.Review.created_at.desc()).all()

    result = []
    for r in reviews:
        user = r.user
        photos = json.loads(r.photos) if r.photos else None
        result.append(schemas.ReviewResponse(
            id=r.id,
            rating=r.rating,
            text=r.text,
            photos=photos,
            venue_id=r.venue_id,
            branch_id=r.branch_id,
            created_at=str(r.created_at),
            user_name=user.name if user else "Удалённый пользователь",
            user_avatar=user.avatar if user else None,
            user_id=r.user_id
        ))
    return result


@app.delete("/api/reviews/{review_id}")
def delete_review(
    review_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    review = db.query(models.Review).filter(
        models.Review.id == review_id,
        models.Review.user_id == current_user.id
    ).first()
    if not review:
        raise HTTPException(status_code=404, detail="Отзыв не найден")
    db.delete(review)
    db.commit()
    return {"message": "Отзыв удален"}


# ==========================================
# АДМИН: 2FA ПОДТВЕРЖДЕНИЕ ВХОДА
# ==========================================

@app.post("/api/admin/verify-2fa")
def verify_admin_2fa(data: schemas.AdminVerifyCode, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == data.email).first()
    if not user or not user.is_admin:
        raise HTTPException(status_code=401, detail="Доступ запрещён")

    stored_code = getattr(user, 'admin_code', None)
    expire_time = getattr(user, 'admin_code_expires', None)
    code_source = "db"

    # Fallback: проверяем in-memory store если колонки не существуют
    if stored_code is None and data.email in _admin_codes:
        fb = _admin_codes[data.email]
        stored_code = fb["code"]
        expire_time = fb["expires"]
        code_source = "memory"

    if not stored_code or stored_code != data.code:
        raise HTTPException(status_code=401, detail="Неверный код подтверждения")

    if expire_time:
        if expire_time.tzinfo is None:
            expire_time = expire_time.replace(tzinfo=timezone.utc)
        if expire_time < datetime.now(timezone.utc):
            raise HTTPException(status_code=401, detail="Код просрочен. Запросите новый.")

    # Очищаем использованный код
    if code_source == "db":
        try:
            user.admin_code = None
            user.admin_code_expires = None
            db.commit()
        except Exception:
            db.rollback()
    else:
        _admin_codes.pop(data.email, None)

    token = auth.create_access_token({"sub": str(user.id)})
    return {
        "access_token": token,
        "token_type": "bearer",
        "user_name": user.name,
        "is_admin": True
    }


@app.post("/api/admin/send-2fa-code")
def resend_admin_2fa(data: schemas.AdminVerifyCode, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == data.email).first()
    if not user or not user.is_admin:
        raise HTTPException(status_code=401, detail="Доступ запрещён")

    admin_code = str(random.randint(1000, 9999))
    code_saved = False
    try:
        user.admin_code = admin_code
        user.admin_code_expires = datetime.now(timezone.utc) + timedelta(minutes=10)
        db.commit()
        code_saved = True
    except Exception:
        db.rollback()

    if not code_saved:
        _admin_codes[user.email] = {
            "code": admin_code,
            "expires": datetime.now(timezone.utc) + timedelta(minutes=10)
        }

    resend_api_key = os.getenv("RESEND_API_KEY")
    if resend_api_key and resend_api_key != "re_СЮДА_ВСТАВЬ_СВОЙ_КЛЮЧ":
        try:
            resend.api_key = resend_api_key
            resend.Emails.send({
                "from": os.getenv("RESEND_FROM_EMAIL", "Minsk Gastro Guide <onboarding@resend.dev>"),
                "to": [user.email],
                "subject": "Новый код подтверждения — Minsk Gastro Guide",
                "text": (
                    f"Здравствуйте, {user.name}!\n\n"
                    f"Ваш новый код для входа: {admin_code}\n\n"
                    f"Код действителен 10 минут."
                )
            })
        except Exception as e:
            print(f"Ошибка отправки 2FA кода: {e}")

    return {"message": "Код отправлен на почту"}


# ==========================================
# АДМИН: ПАНЕЛЬ УПРАВЛЕНИЯ
# ==========================================

@app.get("/api/admin/bookings")
def get_all_bookings(
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(auth.get_current_admin)
):
    bookings = db.query(models.Booking).order_by(models.Booking.created_at.desc()).all()
    return [
        {
            "id": b.id,
            "venue_id": b.venue_id,
            "venue_name": b.venue_name,
            "name": b.name,
            "date": b.date,
            "guests": b.guests,
            "phone": b.phone,
            "message": b.message,
            "cancel_reason": b.cancel_reason,
            "status": b.status,
            "created_at": str(b.created_at)
        }
        for b in bookings
    ]


@app.patch("/api/admin/bookings/{booking_id}/status")
def update_booking_status(
    booking_id: int,
    data: schemas.AdminBookingStatus,
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(auth.get_current_admin)
):
    booking = db.query(models.Booking).filter(models.Booking.id == booking_id).first()
    if not booking:
        raise HTTPException(status_code=404, detail="Бронирование не найдено")

    booking.status = data.status
    if data.reason:
        booking.cancel_reason = data.reason
    db.commit()
    return {"message": "Статус обновлён", "id": booking.id}


@app.get("/api/admin/messages")
def get_all_messages(
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(auth.get_current_admin)
):
    messages = db.query(models.Message).order_by(models.Message.created_at.desc()).all()
    return [
        {
            "id": m.id,
            "name": m.name,
            "email": m.email,
            "subject": m.subject,
            "message": m.message,
            "source": m.source,
            "category": m.category,
            "created_at": str(m.created_at)
        }
        for m in messages
    ]


@app.get("/api/admin/export/bookings")
def export_bookings_csv(
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(auth.get_current_admin)
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import Response
    import io

    bookings = db.query(models.Booking).order_by(models.Booking.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Бронирования"

    headers = ["ID", "Имя", "Телефон", "Заведение", "Филиал (ID)", "Дата", "Гости", "Пожелания", "Статус", "Причина отмены", "Дата создания"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    active_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    cancelled_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    status_font_active = Font(color="2E7D32", bold=True)
    status_font_cancelled = Font(color="C62828", bold=True)

    for row_idx, b in enumerate(bookings, 2):
        status_display = "Активно" if b.status == "active" else "Отменено"
        created_display = b.created_at.strftime("%d.%m.%Y %H:%M") if b.created_at else ""

        values = [b.id, b.name, b.phone, b.venue_name, b.venue_id, b.date, b.guests, b.message or "", status_display, b.cancel_reason or "", created_display]

        row_fill = active_fill if b.status == "active" else cancelled_fill

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col in [8, 10]))

            if col == 9:
                cell.fill = row_fill
                cell.font = status_font_active if b.status == "active" else status_font_cancelled

    col_widths = [6, 20, 18, 25, 15, 14, 8, 30, 12, 25, 18]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bookings_report.xlsx"}
    )


@app.get("/api/admin/export/messages")
def export_messages_csv(
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(auth.get_current_admin)
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import Response
    import io

    messages = db.query(models.Message).order_by(models.Message.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Обращения"

    headers = ["ID", "Имя", "Email", "Тема", "Сообщение", "Источник", "Категория", "Дата"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    contact_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    support_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    source_font_contact = Font(color="1565C0", bold=True)
    source_font_support = Font(color="E65100", bold=True)

    for row_idx, m in enumerate(messages, 2):
        source_display = "Контакты" if m.source == "contact" else "Поддержка"
        created_display = m.created_at.strftime("%d.%m.%Y %H:%M") if m.created_at else ""

        values = [m.id, m.name, m.email, m.subject, m.message, source_display, m.category or "", created_display]

        row_fill = contact_fill if m.source == "contact" else support_fill

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col in [4, 5]))

            if col == 6:
                cell.fill = row_fill
                cell.font = source_font_contact if m.source == "contact" else source_font_support

    col_widths = [6, 20, 25, 30, 40, 14, 18, 18]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=messages_report.xlsx"}
    )


# ==========================================
# КОНТАКТЫ И ПОДДЕРЖКА
# ==========================================

def send_message_email(name: str, email: str, subject: str, body: str, source: str, category: str = None):
    resend_api_key = os.getenv("RESEND_API_KEY")
    print(f"[Resend] Ключ: {'есть' if resend_api_key else 'нет'}")
    if not resend_api_key or resend_api_key == "re_СЮДА_ВСТАВЬ_СВОЙ_КЛЮЧ":
        print("[Resend] Пропускаю отправку — нет API ключа")
        return

    label = "Поддержка" if source == "support" else "Контакты"
    category_line = f"Категория: {category}\n" if category else ""
    to_email = os.getenv("RESEND_TO_EMAIL", "asasin.leha008@gmail.com")

    try:
        resend.api_key = resend_api_key
        result = resend.Emails.send({
            "from": os.getenv("RESEND_FROM_EMAIL", "Minsk Gastro Guide <onboarding@resend.dev>"),
            "to": [to_email],
            "reply_to": [email],
            "subject": f"[{label}] {subject}",
            "text": (
                f"Новое обращение ({label})\n\n"
                f"От: {name}\n"
                f"Email: {email}\n"
                f"Тема: {subject}\n"
                f"{category_line}\n"
                f"Сообщение:\n{body}"
            )
        })
        print(f"[Resend] Письмо отправлено: {result}")
    except Exception as e:
        print(f"[Resend] ОШИБКА: {e}")


@app.post("/api/messages/contact", response_model=schemas.MessageResponse)
def send_contact_message(
    data: schemas.ContactMessageCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user_optional)
):
    user_id = current_user.id if current_user else None
    new_msg = models.Message(
        name=data.name,
        email=data.email,
        subject=data.subject,
        message=data.message,
        source="contact",
        user_id=user_id
    )
    db.add(new_msg)
    db.commit()
    db.refresh(new_msg)

    send_message_email(data.name, data.email, data.subject, data.message, "contact")

    return schemas.MessageResponse(
        id=new_msg.id,
        name=new_msg.name,
        email=new_msg.email,
        subject=new_msg.subject,
        message=new_msg.message,
        source=new_msg.source,
        category=new_msg.category,
        created_at=str(new_msg.created_at)
    )


@app.post("/api/messages/support", response_model=schemas.MessageResponse)
def send_support_message(
    data: schemas.SupportMessageCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    new_msg = models.Message(
        name=data.name,
        email=data.email,
        subject=f"[Поддержка] {data.category}",
        message=data.message,
        source="support",
        category=data.category,
        user_id=current_user.id
    )
    db.add(new_msg)
    db.commit()
    db.refresh(new_msg)

    send_message_email(data.name, data.email, f"[Поддержка] {data.category}", data.message, "support", data.category)

    return schemas.MessageResponse(
        id=new_msg.id,
        name=new_msg.name,
        email=new_msg.email,
        subject=new_msg.subject,
        message=new_msg.message,
        source=new_msg.source,
        category=new_msg.category,
        created_at=str(new_msg.created_at)
    )